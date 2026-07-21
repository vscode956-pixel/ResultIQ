from openpyxl import load_workbook
from pathlib import Path

path = Path('2024-25 BCA Data with Caste (1).xlsx')
print('path exists', path.exists())
wb = load_workbook(path, read_only=True)
print('sheets', wb.sheetnames)
ws = wb.active
headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
print('headers', headers)

# find likely fields
for idx, h in enumerate(headers, start=1):
    print(idx, repr(h))

field_names = {f.lower(): i for i, f in enumerate(headers) if f}
print('field_names', field_names)

religion_col = field_names.get('religion')
category_col = field_names.get('category')
caste_col = field_names.get('caste')
print('religion_col', religion_col, 'category_col', category_col, 'caste_col', caste_col)

if religion_col is not None:
    religions = set()
    for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
        v = row[religion_col]
        if v is not None:
            religions.add(str(v).strip())
    print('sample religions count', len(religions))
    print(sorted(list(religions)))

if category_col is not None:
    categories = set()
    for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
        v = row[category_col]
        if v is not None:
            categories.add(str(v).strip())
    print('sample categories count', len(categories))
    print(sorted(list(categories)))

if caste_col is not None:
    castes = set()
    for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
        v = row[caste_col]
        if v is not None:
            castes.add(str(v).strip())
    print('sample castes count', len(castes))
    print(sorted(list(castes))[:100])
    if len(castes) > 100:
        print('...')
