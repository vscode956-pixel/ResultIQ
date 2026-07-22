import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


HEADER_ALIASES = {
    "USN": [
        "usn",
        "usn no",
        "usn no.",
        "register no",
        "register no.",
        "registration number",
        "registration no",
        "university seat number",
        "universityseatnumber",
        "registerno",
        "registernumber",
    ],
    "StudentName": [
        "student name",
        "studentname",
        "name",
        "candidate name",
        "candidatename",
    ],
    "Gender": [
        "gender",
        "sex",
    ],
    "Category": [
        "category",
        "admission category",
        "admissioncategory",
    ],
    "Caste": [
        "caste",
        "community",
    ],
    "Religion": [
        "religion",
        "religious",
    ],
}

REQUIRED_FIELDS = ["USN", "StudentName"]
OPTIONAL_FIELDS = ["Gender", "Category", "Caste", "Religion"]
ALL_FIELDS = REQUIRED_FIELDS + OPTIONAL_FIELDS


BLANK_ROW_RE = re.compile(r"^[\s\u0000-\u001f\u200b\u200c\u200d\u2060\ufeff]*$")


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    normalized = str(value).strip().lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return normalized.strip()


def matches_alias(header: str, aliases: List[str]) -> bool:
    if not header:
        return False
    header_words = set(header.split())
    for alias in aliases:
        alias_key = normalize_header(alias)
        alias_words = alias_key.split()
        if not alias_words:
            continue
        if header == alias_key:
            return True
        if len(alias_words) == 1:
            word = alias_words[0]
            if word not in header_words:
                continue
            if word == "name" and header_words.intersection({"university", "college", "father", "mother", "parent", "guardian", "equivalent", "institution"}):
                continue
            return True
        if all(word in header_words for word in alias_words):
            return True
    return False


def is_blank_cell(value: Any) -> bool:
    if value is None:
        return True
    text = str(value)
    if BLANK_ROW_RE.match(text):
        return True
    return text.strip() == ""


def detect_header_row(ws, max_scan: int = 30) -> (int, List[str]):
    top_rows = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))

    for row_index, row in enumerate(top_rows, start=1):
        if not row or all(is_blank_cell(cell) for cell in row):
            continue

        raw_headers = [str(cell).strip() if cell is not None else "" for cell in row]
        normalized = [normalize_header(cell) for cell in raw_headers]

        recognized = set()
        for field, aliases in HEADER_ALIASES.items():
            for name in normalized:
                if matches_alias(name, aliases):
                    recognized.add(field)
                    break

        if "USN" in recognized and "StudentName" in recognized:
            return row_index, raw_headers

    for row_index, row in enumerate(top_rows, start=1):
        if row and any(not is_blank_cell(cell) for cell in row):
            raw_headers = [str(cell).strip() if cell is not None else "" for cell in row]
            return row_index, raw_headers

    return 1, []


def build_mapping(header_row: List[str]) -> (Dict[str, Dict[str, Optional[Any]]], List[str]):
    normalized = [normalize_header(cell) for cell in header_row]
    mapping: Dict[str, Dict[str, Optional[Any]]] = {}
    missing: List[str] = []

    for field, aliases in HEADER_ALIASES.items():
        found_index = None
        found_header = None
        for index, header in enumerate(normalized):
            if not header:
                continue
            if matches_alias(header, aliases):
                found_index = index
                found_header = header_row[index]
                break
        if found_index is None:
            missing.append(field)
        mapping[field] = {"index": found_index, "header": found_header}

    return mapping, missing


def parse_students(excel_path: str) -> Dict[str, Any]:
    wb = load_workbook(filename=excel_path, data_only=True, read_only=True)
    ws = wb.active

    header_row_index, header_row = detect_header_row(ws, max_scan=30)
    mapping, missing_fields = build_mapping(header_row)

    required_missing = [field for field in REQUIRED_FIELDS if mapping[field]["index"] is None]
    if required_missing:
        return {
            "error": True,
            "message": f"Missing required columns: {', '.join(required_missing)}",
            "header_row_index": header_row_index,
            "mapping": mapping,
            "rows": [],
            "missing_counts": {},
            "warnings": [],
        }

    optional_missing = [field for field in OPTIONAL_FIELDS if mapping[field]["index"] is None]
    rows: List[Dict[str, Optional[str]]] = []
    missing_counts = {field: 0 for field in REQUIRED_FIELDS}
    warnings: List[str] = []

    if optional_missing:
        warnings.append(f"Optional columns not found: {', '.join(optional_missing)}")

    for row_index, row in enumerate(ws.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
        if row is None or all(is_blank_cell(cell) for cell in row):
            continue

        values = [cell for cell in row]
        student: Dict[str, Optional[str]] = {}
        row_has_data = False

        for field in ALL_FIELDS:
            col_index = mapping[field]["index"]
            value = None
            if col_index is not None and col_index < len(values):
                raw_value = values[col_index]
                if raw_value is not None and str(raw_value).strip() != "":
                    value = str(raw_value).strip()
                    row_has_data = True
            student[field] = value

        if not row_has_data:
            continue

        if not student["USN"]:
            warnings.append(f"Missing USN at row {row_index}")
            missing_counts["USN"] += 1

        if not student["StudentName"]:
            warnings.append(f"Missing Name at row {row_index}")
            missing_counts["StudentName"] += 1

        for field in OPTIONAL_FIELDS:
            if mapping[field]["index"] is not None and student[field] is None:
                warnings.append(f"Missing {field} at row {row_index}")

        rows.append(student)

    return {
        "error": False,
        "header_row_index": header_row_index,
        "mapping": mapping,
        "rows": rows,
        "missing_counts": missing_counts,
        "warnings": warnings,
    }


def print_results(file_path: str, parsed: Dict[str, Any]) -> None:
    file_name = Path(file_path).name
    print("=" * 49)
    print("          EXCEL PARSER TEST")
    print("=" * 49)
    print()
    print("Opening File:")
    print(file_name)
    print()

    if parsed["error"]:
        print(f"Error: {parsed['message']}")
        return

    header_row_index = parsed["header_row_index"]
    mapping = parsed["mapping"]
    rows = parsed["rows"]
    missing_counts = parsed["missing_counts"]
    warnings = parsed["warnings"]

    print(f"✓ Header row detected at Row {header_row_index}")
    print()
    print("Detected Columns")
    print("--------------------------------")
    for field in REQUIRED_FIELDS:
        info = mapping[field]
        idx = info["index"]
        column_name = get_column_letter(idx + 1) if idx is not None else "?"
        display_name = field.replace("StudentName", "Student Name")
        print(f"{display_name:<14} -> Column {column_name}")
    print()

    for index, student in enumerate(rows, start=1):
        print("=" * 46)
        print(f"Student {index}")
        print("----------------------------------------------")
        print(f"USN      : {student.get('USN') or ''}")
        print(f"Name     : {student.get('StudentName') or ''}")
        print(f"Gender   : {student.get('Gender') or ''}")
        print(f"Category : {student.get('Category') or ''}")
        print(f"Caste    : {student.get('Caste') or ''}")
        print()

    print("=" * 49)
    print("SUMMARY")
    print("=" * 49)
    print()
    print(f"Total Students : {len(rows)}")
    print()
    print(f"Missing USN      : {missing_counts['USN']}")
    print(f"Missing Name     : {missing_counts['StudentName']}")
    print(f"Missing Gender   : {missing_counts['Gender']}")
    print(f"Missing Category : {missing_counts['Category']}")
    print(f"Missing Caste    : {missing_counts['Caste']}")

    if warnings:
        print()
        print("Warnings:")
        for warning in warnings:
            print(f"- {warning}")

    print()
    print("Excel Parsing Successful")


def main() -> None:
    file_path = sys.argv[1] if len(sys.argv) > 1 else "sample.xlsx"
    if not Path(file_path).exists():
        print(f"File not found: {file_path}")
        sys.exit(1)

    parsed = parse_students(file_path)
    print_results(file_path, parsed)


if __name__ == "__main__":
    main()
